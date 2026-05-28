from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app import crud, schemas
from app.database import get_db


router = APIRouter(prefix="/users", tags=["Usuarios"])


@router.get("", response_model=list[schemas.UserRead])
def get_users(db: Session = Depends(get_db)) -> list[schemas.UserRead]:
    return crud.list_users(db)


@router.post("", response_model=schemas.UserRead, status_code=201)
def post_user(payload: schemas.UserCreate, db: Session = Depends(get_db)) -> schemas.UserRead:
    return crud.create_user(db, payload)


@router.get("/bulk-template")
def download_bulk_template() -> StreamingResponse:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "usuarios"
    headers = ["first_name", "last_name", "email", "document_number", "role", "curso", "materia", "acudiente_id"]
    sheet.append(headers)
    sheet.append(["Ana", "Lopez", "ana.lopez@colegiosol.edu.co", "10980001", "STUDENT", "", "", "UUID_ACUDIENTE"])
    sheet.append(["Juan", "Perez", "juan.perez@colegiosol.edu.co", "80112233", "TEACHER", "Primero A", "Matematicas", ""])

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="plantilla_usuarios_colegio_sol.xlsx"'},
    )


@router.post("/bulk-upload", response_model=schemas.BulkUploadResult)
async def bulk_upload_users(file: UploadFile, db: Session = Depends(get_db)) -> schemas.BulkUploadResult:
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="El archivo debe ser .xlsx.")

    content = await file.read()
    try:
        workbook = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="No fue posible leer el archivo Excel.") from exc

    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required = crud.REQUIRED_BULK_COLUMNS
    missing = sorted(required - set(headers))
    if missing:
        raise HTTPException(status_code=400, detail=f"Faltan columnas obligatorias: {', '.join(missing)}.")

    rows: list[dict[str, str]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rows.append({header: "" if value is None else str(value) for header, value in zip(headers, row)})

    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no contiene usuarios para importar.")

    return crud.bulk_create_users(db, rows)
